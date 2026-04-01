import openpyxl
import csv
import io

from rest_framework.viewsets import ModelViewSet
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.views import APIView

from django.contrib.auth import get_user_model
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone

from .models import (
    Lead,
    CampaignLead,
    ThirdPartyLead,
    InternalLead, LeadAssignmentHistory,
    OnlineLead, OnlineLeadAssignmentHistory,
    UsedLead, UsedLeadAssignmentHistory,
    ArchivedLead, ArchivedLeadAssignmentHistory,
    UploadData, AllocateData, ReallocateAssignedLead,
)
from .serializers import (
    LeadSerializer,
    CampaignLeadSerializer,
    ThirdPartyLeadSerializer,
    InternalLeadSerializer,
    OnlineLeadSerializer,
    UsedLeadSerializer,
    ArchivedLeadSerializer,
    ReallocateAssignedLeadSerializer,
)
from data_lead.services import fetch_from_api, process_third_party_leads

User = get_user_model()


# ─────────────────────────────────────────────────────────────────────────────
# Lead
# ─────────────────────────────────────────────────────────────────────────────

class LeadViewSet(ModelViewSet):
    queryset         = Lead.objects.all()
    serializer_class = LeadSerializer


# ─────────────────────────────────────────────────────────────────────────────
# Users list (used by frontend agent dropdowns)
# ─────────────────────────────────────────────────────────────────────────────

class UserListView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        users = User.objects.values("id", "username")
        return Response(list(users))


# ─────────────────────────────────────────────────────────────────────────────
# CampaignLead
# ─────────────────────────────────────────────────────────────────────────────

class CampaignLeadViewSet(ModelViewSet):
    queryset           = CampaignLead.objects.all().order_by("-created_at")
    serializer_class   = CampaignLeadSerializer
    parser_classes     = [MultiPartParser, FormParser, JSONParser]
    permission_classes = [AllowAny]
    authentication_classes = []

    def perform_create(self, serializer):
        serializer.save()

    @action(detail=False, methods=["get"])
    def unassigned(self, request):
        qs = CampaignLead.objects.filter(assigned_users__isnull=True)
        return Response(self.get_serializer(qs, many=True).data)

    @action(detail=False, methods=["post"], url_path="upload-file")
    def upload_file(self, request):
        import sys
        print("=== upload_file called ===", file=sys.stderr)

        file = request.FILES.get("file") or (
            next(iter(request.FILES.values()), None) if request.FILES else None
        )
        if not file:
            return Response(
                {"error": "No file provided.",
                 "debug": {"files_keys": list(request.FILES.keys()),
                           "data_keys":  list(request.data.keys()),
                           "content_type": request.content_type}},
                status=status.HTTP_400_BAD_REQUEST,
            )

        filename = file.name.lower()
        rows = []
        try:
            if filename.endswith((".xlsx", ".xls")):
                wb      = openpyxl.load_workbook(file, data_only=True)
                ws      = wb.active
                headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(v is not None for v in row):
                        rows.append(dict(zip(headers, row)))
            elif filename.endswith(".csv"):
                decoded = file.read().decode("utf-8-sig")
                rows    = list(csv.DictReader(io.StringIO(decoded)))
            else:
                return Response(
                    {"error": f"Unsupported file type '{filename}'. Use CSV or XLSX."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except Exception as e:
            return Response({"error": f"Failed to parse file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        created, failed = [], []
        for i, row in enumerate(rows, start=2):
            raw_consent          = str(row.get("consent_obtained", "")).strip().lower()
            row["consent_obtained"] = raw_consent in ("true", "1", "yes")
            cleaned = {k: (str(v).strip() if v is not None else "") for k, v in row.items()}
            cleaned["consent_obtained"] = row["consent_obtained"]
            for opt in ["notes", "follow_up_date", "campaign_end", "conversion_status", "tags"]:
                if cleaned.get(opt) == "":
                    cleaned.pop(opt, None)
            serializer = self.get_serializer(data=cleaned)
            if serializer.is_valid():
                serializer.save()
                created.append(serializer.data)
            else:
                failed.append({"row": i, "data": cleaned, "errors": serializer.errors})

        return Response(
            {"message": f"{len(created)} leads created, {len(failed)} failed.",
             "created_count": len(created), "failed_count": len(failed), "failed_rows": failed},
            status=status.HTTP_201_CREATED if created else status.HTTP_400_BAD_REQUEST,
        )

    @action(detail=False, methods=["post"])
    def allocate(self, request):
        user_id  = request.data.get("user_id")
        lead_ids = request.data.get("lead_ids", [])
        if not user_id or not lead_ids:
            return Response({"error": "user_id and lead_ids are required"}, status=status.HTTP_400_BAD_REQUEST)
        try:
            user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response({"error": f"User {user_id} not found"}, status=status.HTTP_404_NOT_FOUND)
        leads = CampaignLead.objects.filter(id__in=lead_ids)
        for lead in leads:
            lead.assigned_users.add(user)
        return Response({"message": f"{leads.count()} leads allocated successfully"})

    @action(detail=False, methods=["post"])
    def reallocate(self, request):
        from_user_id = request.data.get("from_user_id")
        to_user_id   = request.data.get("to_user_id")
        lead_ids     = request.data.get("lead_ids", [])
        if not to_user_id or not lead_ids:
            return Response({"error": "to_user_id and lead_ids are required"}, status=status.HTTP_400_BAD_REQUEST)
        try:
            to_user = User.objects.get(id=to_user_id)
        except User.DoesNotExist:
            return Response({"error": f"User {to_user_id} not found"}, status=status.HTTP_404_NOT_FOUND)
        leads = CampaignLead.objects.filter(id__in=lead_ids)
        for lead in leads:
            if from_user_id:
                try:
                    lead.assigned_users.remove(User.objects.get(id=from_user_id))
                except User.DoesNotExist:
                    pass
            else:
                lead.assigned_users.clear()
            lead.assigned_users.add(to_user)
        return Response({"message": f"{leads.count()} leads reallocated successfully"})


# ─────────────────────────────────────────────────────────────────────────────
# ThirdPartyLead
# ─────────────────────────────────────────────────────────────────────────────

class ThirdPartyLeadViewSet(ModelViewSet):
    queryset         = ThirdPartyLead.objects.all().order_by("-created_at")
    serializer_class = ThirdPartyLeadSerializer
    parser_classes   = [MultiPartParser, FormParser]

    def perform_create(self, serializer):
        upload = serializer.save(status="PENDING")
        process_third_party_leads(upload)

    @action(detail=False, methods=["post"], url_path="manual")
    def manual_entry(self, request):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save(source_type="manual")
            return Response({"message": "Lead created successfully", "data": serializer.data}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["post"], url_path="api-sync")
    def api_sync(self, request):
        leads = request.data.get("leads")
        if not isinstance(leads, list):
            return Response({"error": "Expected 'leads' to be a list"}, status=status.HTTP_400_BAD_REQUEST)
        created, failed = [], []
        for lead_data in leads:
            serializer = self.get_serializer(data=lead_data)
            if serializer.is_valid():
                serializer.save(source_type="api")
                created.append(serializer.data)
            else:
                failed.append({"data": lead_data, "errors": serializer.errors})
        return Response({"created_count": len(created), "failed_count": len(failed), "created": created, "failed": failed}, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=["post"], url_path="upload-file")
    def upload_file(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"error": "CSV file is required"}, status=status.HTTP_400_BAD_REQUEST)
        if not file.name.endswith(".csv"):
            return Response({"error": "Only CSV files are supported"}, status=status.HTTP_400_BAD_REQUEST)
        reader  = csv.DictReader(io.StringIO(file.read().decode("utf-8")))
        created, failed = [], []
        for row in reader:
            serializer = self.get_serializer(data=row)
            if serializer.is_valid():
                serializer.save(source_type="file")
                created.append(serializer.data)
            else:
                failed.append({"row": row, "errors": serializer.errors})
        return Response({"created_count": len(created), "failed_count": len(failed), "created": created, "failed": failed}, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=["get"])
    def unassigned(self, request):
        qs = ThirdPartyLead.objects.filter(assigned_users__isnull=True)
        return Response(self.get_serializer(qs, many=True).data)

    @action(detail=False, methods=["get"])
    def assigned(self, request):
        user_id = request.query_params.get("user_id")
        qs = ThirdPartyLead.objects.filter(assigned_users__isnull=False)
        if user_id:
            qs = qs.filter(assigned_users__id=user_id)
        return Response(self.get_serializer(qs, many=True).data)

    @action(detail=False, methods=["post"])
    def allocate(self, request):
        user_id  = request.data.get("user_id")
        lead_ids = request.data.get("lead_ids", [])
        if not user_id or not lead_ids:
            return Response({"error": "user_id and lead_ids are required"}, status=status.HTTP_400_BAD_REQUEST)
        user  = User.objects.get(id=user_id)
        leads = ThirdPartyLead.objects.filter(id__in=lead_ids, assigned_users__isnull=True)
        for lead in leads:
            lead.assigned_users.add(user)
        return Response({"message": f"{leads.count()} leads allocated successfully"}, status=status.HTTP_200_OK)

    @action(detail=False, methods=["post"])
    def reallocate(self, request):
        from_user_id = request.data.get("from_user_id")
        to_user_id   = request.data.get("to_user_id")
        lead_ids     = request.data.get("lead_ids", [])
        if not to_user_id or not lead_ids:
            return Response({"error": "to_user_id and lead_ids are required"}, status=status.HTTP_400_BAD_REQUEST)
        to_user = User.objects.get(id=to_user_id)
        leads   = ThirdPartyLead.objects.filter(id__in=lead_ids)
        for lead in leads:
            if from_user_id:
                lead.assigned_users.remove(from_user_id)
            else:
                lead.assigned_users.clear()
            lead.assigned_users.add(to_user)
        return Response({"message": f"{leads.count()} leads reallocated successfully"}, status=status.HTTP_200_OK)


# ─────────────────────────────────────────────────────────────────────────────
# InternalLead
# ─────────────────────────────────────────────────────────────────────────────

class InternalLeadViewSet(ModelViewSet):
    queryset           = InternalLead.objects.all().order_by("-created_at")
    serializer_class   = InternalLeadSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        qs     = super().get_queryset()
        params = self.request.query_params
        if params.get("department"):
            qs = qs.filter(referrer_department_id=params["department"])
        if params.get("status"):
            qs = qs.filter(lead_status=params["status"])
        if params.get("quality"):
            qs = qs.filter(lead_quality=params["quality"])
        if params.get("source"):
            qs = qs.filter(internal_source=params["source"])
        if params.get("assigned_to"):
            qs = qs.filter(assigned_to_id=params["assigned_to"])
        if params.get("high_priority") == "true":
            qs = qs.filter(high_priority=True)
        if params.get("search"):
            qs = qs.filter(contact_name__icontains=params["search"])
        return qs

    @action(detail=False, methods=["post"], url_path="employee-referral")
    def employee_referral(self, request):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save(Upload_method="employee_refferal")
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["post"], url_path="manual")
    def manual_entry(self, request):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save(Upload_method="manual")
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["post"], url_path="internal-api")
    def internal_api(self, request):
        leads          = request.data.get("leads", [])
        created, failed = [], []
        for lead in leads:
            serializer = self.get_serializer(data=lead)
            if serializer.is_valid():
                serializer.save(Upload_method="internal_api")
                created.append(serializer.data)
            else:
                failed.append(serializer.errors)
        return Response({"created": len(created), "failed": len(failed), "errors": failed}, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=["post"], url_path="allocate")
    def allocate(self, request):
        lead_ids = request.data.get("lead_ids", [])
        agent_id = request.data.get("agent_id")
        if not lead_ids or not agent_id:
            return Response({"error": "lead_ids and agent_id required"}, status=status.HTTP_400_BAD_REQUEST)
        leads      = InternalLead.objects.filter(id__in=lead_ids)
        changed_by = request.user if request.user.is_authenticated else None
        updated    = []
        for lead in leads:
            old_owner         = lead.assigned_to
            lead.assigned_to_id = agent_id
            lead.save()
            LeadAssignmentHistory.objects.create(lead=lead, from_user=old_owner, to_user_id=agent_id, changed_by=changed_by, note="Bulk allocation")
            updated.append(lead.id)
        return Response({"message": "Leads allocated successfully", "updated_count": len(updated)})

    @action(detail=True, methods=["post"], url_path="reallocate")
    def reallocate(self, request, pk=None):
        lead         = self.get_object()
        new_owner_id = request.data.get("agent_id")
        note         = request.data.get("note", "")
        if not new_owner_id:
            return Response({"error": "agent_id required"}, status=status.HTTP_400_BAD_REQUEST)
        old_owner           = lead.assigned_to
        lead.assigned_to_id = new_owner_id
        lead.save()
        changed_by = request.user if request.user.is_authenticated else None
        LeadAssignmentHistory.objects.create(lead=lead, from_user=old_owner, to_user_id=new_owner_id, changed_by=changed_by, note=note)
        return Response({"message": "Lead reallocated successfully"})

    @action(detail=False, methods=["post"], url_path="upload-file")
    def upload_file(self, request):
        file = request.FILES.get("file") or (next(iter(request.FILES.values()), None) if request.FILES else None)
        if not file:
            return Response({"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)
        filename = file.name.lower()
        rows = []
        try:
            if filename.endswith((".xlsx", ".xls")):
                wb      = openpyxl.load_workbook(file, data_only=True)
                ws      = wb.active
                headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(v is not None for v in row):
                        rows.append(dict(zip(headers, row)))
            elif filename.endswith(".csv"):
                decoded = file.read().decode("utf-8-sig")
                rows    = list(csv.DictReader(io.StringIO(decoded)))
            else:
                return Response({"error": "Use CSV or XLSX."}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": f"Parse error: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
        created, failed = [], []
        for i, row in enumerate(rows, start=2):
            raw_consent          = str(row.get("consent_obtained", "")).strip().lower()
            row["consent_obtained"] = raw_consent in ("true", "1", "yes")
            cleaned = {k: (str(v).strip() if v is not None else "") for k, v in row.items()}
            cleaned["consent_obtained"] = row["consent_obtained"]
            cleaned["Upload_method"]    = "file"
            for opt in ["notes", "follow_up_date", "tags", "internal_lead_id", "api_endpoint", "api_key"]:
                if cleaned.get(opt) == "":
                    cleaned.pop(opt, None)
            serializer = self.get_serializer(data=cleaned)
            if serializer.is_valid():
                serializer.save()
                created.append(serializer.data)
            else:
                failed.append({"row": i, "errors": serializer.errors})
        return Response({"message": f"{len(created)} leads created, {len(failed)} failed.", "created_count": len(created), "failed_count": len(failed), "failed_rows": failed},
                        status=status.HTTP_201_CREATED if created else status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["get"], url_path="stats")
    def stats(self, request):
        from django.db.models import Count
        qs     = InternalLead.objects.all()
        total  = qs.count()
        now    = timezone.now()
        active = qs.filter(created_at__year=now.year, created_at__month=now.month).count()
        high   = qs.filter(lead_quality="HIGH").count()
        hot    = qs.filter(lead_status="HOT").count()
        rate   = round((hot / total * 100), 1) if total else 0
        top    = qs.values("internal_source").annotate(count=Count("id")).order_by("-count").first()
        return Response({"total_internal_leads": total, "active_this_month": active, "high_quality_leads": high,
                         "conversion_rate": rate, "top_referrer": top["internal_source"] if top else "N/A", "top_referrer_count": top["count"] if top else 0})


# ─────────────────────────────────────────────────────────────────────────────
# OnlineLead
# ─────────────────────────────────────────────────────────────────────────────

class OnlineLeadViewSet(ModelViewSet):
    queryset           = OnlineLead.objects.all().order_by("-created_at")
    serializer_class   = OnlineLeadSerializer
    permission_classes = [AllowAny]
    authentication_classes = []
    parser_classes     = [MultiPartParser, FormParser, JSONParser]

    def get_queryset(self):
        qs     = super().get_queryset()
        params = self.request.query_params
        if params.get("search"):    qs = qs.filter(contact_name__icontains=params["search"])
        if params.get("status"):    qs = qs.filter(lead_status=params["status"])
        if params.get("quality"):   qs = qs.filter(lead_quality=params["quality"])
        if params.get("source"):    qs = qs.filter(online_source__icontains=params["source"])
        if params.get("assigned_to"): qs = qs.filter(assigned_to_id=params["assigned_to"])
        if params.get("unassigned") == "true": qs = qs.filter(assigned_to__isnull=True)
        if params.get("assigned")   == "true": qs = qs.filter(assigned_to__isnull=False)
        return qs

    @action(detail=False, methods=["get"], url_path="stats")
    def stats(self, request):
        from django.db.models import Count
        qs    = OnlineLead.objects.all()
        total = qs.count()
        now   = timezone.now()
        active = qs.filter(created_at__year=now.year, created_at__month=now.month).count()
        high   = qs.filter(lead_quality="HIGH").count()
        hot    = qs.filter(lead_status="HOT").count()
        rate   = round((hot / total * 100), 1) if total else 0
        top    = qs.values("online_source").annotate(count=Count("id")).order_by("-count").first()
        return Response({"total_online_leads": total, "active_this_month": active, "high_quality_leads": high,
                         "conversion_rate": rate, "top_source": top["online_source"] if top else "N/A", "top_source_count": top["count"] if top else 0})

    @action(detail=False, methods=["post"], url_path="upload-file")
    def upload_file(self, request):
        file = request.FILES.get("file") or (next(iter(request.FILES.values()), None) if request.FILES else None)
        if not file:
            return Response({"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)
        filename = file.name.lower()
        rows = []
        try:
            if filename.endswith((".xlsx", ".xls")):
                wb = openpyxl.load_workbook(file, data_only=True)
                ws = wb.active
                headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(v is not None for v in row): rows.append(dict(zip(headers, row)))
            elif filename.endswith(".csv"):
                decoded = file.read().decode("utf-8-sig")
                rows = list(csv.DictReader(io.StringIO(decoded)))
            else:
                return Response({"error": "Use CSV or XLSX."}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": f"Parse error: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
        created, failed = [], []
        for i, row in enumerate(rows, start=2):
            raw = str(row.get("consent_obtained", "")).strip().lower()
            row["consent_obtained"] = raw in ("true", "1", "yes")
            cleaned = {k: (str(v).strip() if v is not None else "") for k, v in row.items()}
            cleaned["consent_obtained"] = row["consent_obtained"]
            for opt in ["notes", "follow_up_date", "tags", "online_lead_id", "lead_form_url", "api_endpoint", "api_key"]:
                if cleaned.get(opt) == "": cleaned.pop(opt, None)
            serializer = self.get_serializer(data=cleaned)
            if serializer.is_valid():
                serializer.save()
                created.append(serializer.data)
            else:
                failed.append({"row": i, "errors": serializer.errors})
        return Response({"message": f"{len(created)} leads created, {len(failed)} failed.", "created_count": len(created), "failed_count": len(failed), "failed_rows": failed},
                        status=status.HTTP_201_CREATED if created else status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["post"], url_path="allocate")
    def allocate(self, request):
        lead_ids = request.data.get("lead_ids", [])
        agent_id = request.data.get("agent_id")
        if not lead_ids or not agent_id:
            return Response({"error": "lead_ids and agent_id are required"}, status=status.HTTP_400_BAD_REQUEST)
        leads      = OnlineLead.objects.filter(id__in=lead_ids)
        changed_by = request.user if request.user.is_authenticated else None
        updated    = 0
        for lead in leads:
            old = lead.assigned_to
            lead.assigned_to_id = agent_id
            lead.save()
            OnlineLeadAssignmentHistory.objects.create(lead=lead, from_user=old, to_user_id=agent_id, changed_by=changed_by, note="Bulk allocation")
            updated += 1
        return Response({"message": "Online leads allocated successfully", "updated_count": updated})

    @action(detail=True, methods=["post"], url_path="reallocate")
    def reallocate(self, request, pk=None):
        lead         = self.get_object()
        new_owner_id = request.data.get("agent_id")
        note         = request.data.get("note", "")
        if not new_owner_id:
            return Response({"error": "agent_id is required"}, status=status.HTTP_400_BAD_REQUEST)
        old = lead.assigned_to
        lead.assigned_to_id = new_owner_id
        lead.save()
        changed_by = request.user if request.user.is_authenticated else None
        OnlineLeadAssignmentHistory.objects.create(lead=lead, from_user=old, to_user_id=new_owner_id, changed_by=changed_by, note=note)
        return Response({"message": "Lead reallocated successfully"})

class UsedLeadViewSet(viewsets.ModelViewSet):
    serializer_class   = UsedLeadSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        qs = (
            UsedLead.objects
            .select_related("product", "allocated_to")
            .prefetch_related("lead_object")   # remove this if GenericFK causes crash
            .order_by("-created_at")
        )

        p = self.request.query_params

        if p.get("source"):
            qs = qs.filter(Source=p["source"])
        if p.get("agent"):
            qs = qs.filter(allocated_to_id=p["agent"])
        if p.get("status"):
            qs = qs.filter(status=p["status"])
        if p.get("outcome"):
            qs = qs.filter(outcome=p["outcome"])
        if p.get("product"):
            qs = qs.filter(product_id=p["product"])

        is_active = p.get("is_active")
        if is_active is not None:
            qs = qs.filter(is_active=is_active.lower() == "true")

        return qs


    @action(detail=False, methods=["get"], url_path="all_data")
    def all_data(self, request):
        try:
            qs = self.get_queryset()
            page = self.paginate_queryset(qs)

            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(serializer.data)

            serializer = self.get_serializer(qs, many=True)
            return Response(serializer.data)

        except Exception as e:
            return Response({"error": str(e)}, status=500)


    @action(detail=False, methods=["post"], url_path="allocate")
    def allocate(self, request):
        lead_ids   = request.data.get("lead_ids", [])
        agent_id   = request.data.get("agent_id")
        product_id = request.data.get("product_id")

        if not lead_ids or not agent_id:
            return Response(
                {"error": "lead_ids and agent_id are required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            agent = User.objects.get(id=agent_id)
        except User.DoesNotExist:
            return Response(
                {"error": f"User {agent_id} not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        qs = UsedLead.objects.filter(id__in=lead_ids)
        if product_id:
            qs = qs.filter(product_id=product_id)

        updated = 0
        changed_by = request.user if request.user.is_authenticated else None

        for lead in qs:
            old_agent = lead.allocated_to
            lead.allocated_to = agent
            lead.save(update_fields=["allocated_to"])

            UsedLeadAssignmentHistory.objects.create(
                lead=lead,
                from_agent=old_agent,
                to_agent=agent,
                changed_by=changed_by,
                reason="Initial allocation"
            )
            updated += 1

        return Response({
            "message": f"{updated} lead(s) allocated successfully.",
            "updated_count": updated
        })


    @action(detail=True, methods=["post"], url_path="reallocate")
    def reallocate(self, request, pk=None):
        lead     = self.get_object()
        agent_id = request.data.get("agent_id")
        reason   = request.data.get("reason", "").strip()

        if not agent_id or not reason:
            return Response(
                {"error": "agent_id and reason are required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            new_agent = User.objects.get(id=agent_id)
        except User.DoesNotExist:
            return Response(
                {"error": f"User {agent_id} not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        old_agent = lead.allocated_to
        lead.allocated_to = new_agent
        lead.save(update_fields=["allocated_to"])

        changed_by = request.user if request.user.is_authenticated else None

        UsedLeadAssignmentHistory.objects.create(
            lead=lead,
            from_agent=old_agent,
            to_agent=new_agent,
            changed_by=changed_by,
            reason=reason
        )

        return Response({"message": "Lead reallocated successfully."})
# ─────────────────────────────────────────────────────────────────────────────
# ArchivedLead
# ─────────────────────────────────────────────────────────────────────────────

class ArchivedLeadViewSet(ModelViewSet):
    queryset           = ArchivedLead.objects.all().order_by("-archived_at")
    serializer_class   = ArchivedLeadSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        qs = super().get_queryset()
        p  = self.request.query_params
        if p.get("assigned_to"):   qs = qs.filter(assigned_to_id=p["assigned_to"])
        if p.get("archive_reason"): qs = qs.filter(archive_reason=p["archive_reason"])
        if p.get("start_date") and p.get("end_date"):
            qs = qs.filter(archived_at__date__range=[p["start_date"], p["end_date"]])
        return qs

    def update(self, request, *args, **kwargs):
        return Response({"detail": "Archived leads cannot be edited."}, status=status.HTTP_403_FORBIDDEN)

    def destroy(self, request, *args, **kwargs):
        return Response({"detail": "Archived leads cannot be deleted."}, status=status.HTTP_403_FORBIDDEN)

    def perform_create(self, serializer):
        serializer.save(archived_by=self.request.user)

    @action(detail=False, methods=["post"], url_path="allocate")
    def allocate(self, request):
        lead_ids = request.data.get("lead_ids", [])
        user_id  = request.data.get("user_id")
        if not lead_ids or not user_id:
            return Response({"error": "lead_ids and user_id required"}, status=status.HTTP_400_BAD_REQUEST)
        leads   = ArchivedLead.objects.filter(id__in=lead_ids)
        updated = 0
        for lead in leads:
            old = lead.assigned_to
            lead.assigned_to_id = user_id
            lead.save()
            ArchivedLeadAssignmentHistory.objects.create(lead=lead, from_user=old, to_user_id=user_id, changed_by=request.user, reason="Allocation on archived lead")
            updated += 1
        return Response({"message": "Archived leads allocated successfully", "updated_count": updated})

    @action(detail=True, methods=["post"], url_path="reallocate")
    def reallocate(self, request, pk=None):
        lead       = self.get_object()
        new_user_id = request.data.get("user_id")
        reason     = request.data.get("reason")
        if not new_user_id or not reason:
            return Response({"error": "user_id and reason are required"}, status=status.HTTP_400_BAD_REQUEST)
        old = lead.assigned_to
        lead.assigned_to_id = new_user_id
        lead.save()
        ArchivedLeadAssignmentHistory.objects.create(lead=lead, from_user=old, to_user_id=new_user_id, changed_by=request.user, reason=reason)
        return Response({"message": "Archived lead reallocated successfully"})

    @action(detail=True, methods=["post"])
    def reactivate(self, request, pk=None):
        lead = self.get_object()
        if not request.user.is_staff:
            return Response({"detail": "Permission denied"}, status=status.HTTP_403_FORBIDDEN)
        lead.is_reactivated  = True
        lead.reactivated_by  = request.user
        lead.reactivated_at  = timezone.now()
        lead.save()
        return Response({"message": "Lead reactivated successfully"})

    @action(detail=False, methods=["get"])
    def all_data(self, request):
        is_reactivated = request.query_params.get("is_reactivated")
        qs = self.get_queryset()
        if is_reactivated is not None:
            qs = qs.filter(is_reactivated=is_reactivated.lower() == "true")
        return Response(self.get_serializer(qs, many=True).data)

    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        from django.http import HttpResponse
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="archived_leads.csv"'
        writer = csv.writer(response)
        writer.writerow(["Customer Name", "Phone", "Email", "Product", "Assigned To", "Archive Reason", "Archived At"])
        for lead in self.get_queryset():
            writer.writerow([
                getattr(lead, "customer_name", ""),
                getattr(lead, "phone", ""),
                getattr(lead, "email", ""),
                lead.product.name if lead.product else "",
                lead.assigned_to.get_full_name() if lead.assigned_to else "",
                getattr(lead, "archive_reason", lead.archived_reason),
                lead.archived_at,
            ])
        return response


# ─────────────────────────────────────────────────────────────────────────────
# UploadData / AllocateData (template views)
# ─────────────────────────────────────────────────────────────────────────────

def upload_data_list(request):
    return render(request, "upload_data/list.html", {"data": UploadData.objects.all().order_by("-created_at")})

def upload_data_create(request):
    if request.method == "POST":
        UploadData.objects.create(
            configuration_name=request.POST.get("configuration_name"),
            file_upload=request.FILES.get("file_upload"),
            product=request.POST.get("product"),
            lead_status=request.POST.get("lead_status"),
            lead_source=request.POST.get("lead_source"),
            contact_name=request.POST.get("contact_name"),
            mobile_number=request.POST.get("mobile_number"),
            email=request.POST.get("email"),
            assigned_user=request.POST.get("assigned_user"),
            conversion_status=request.POST.get("conversion_status"),
            campaign_start_date=request.POST.get("campaign_start_date"),
            campaign_end_date=request.POST.get("campaign_end_date"),
            follow_up_date=request.POST.get("follow_up_date"),
            notes=request.POST.get("notes"),
            tags=request.POST.get("tags"),
            consent_obtained=bool(request.POST.get("consent_obtained")),
        )
        messages.success(request, "Data Created Successfully")
        return redirect("upload_data_list")
    return render(request, "upload_data/create.html")

def upload_data_update(request, pk):
    data = get_object_or_404(UploadData, pk=pk)
    if request.method == "POST":
        for field in ["configuration_name", "product", "lead_status", "lead_source",
                      "contact_name", "mobile_number", "email", "assigned_user",
                      "conversion_status", "campaign_start_date", "campaign_end_date",
                      "follow_up_date", "notes", "tags"]:
            setattr(data, field, request.POST.get(field))
        data.consent_obtained = bool(request.POST.get("consent_obtained"))
        if request.FILES.get("file_upload"):
            data.file_upload = request.FILES["file_upload"]
        data.save()
        messages.success(request, "Data Updated Successfully")
        return redirect("upload_data_list")
    return render(request, "upload_data/update.html", {"data": data})

def upload_data_delete(request, pk):
    get_object_or_404(UploadData, pk=pk).delete()
    messages.success(request, "Data Deleted Successfully")
    return redirect("upload_data_list")

def allocate_data_list(request):
    return render(request, "allocate_data/list.html", {"data": AllocateData.objects.all().order_by("-created_at")})

def allocate_data_create(request):
    if request.method == "POST":
        AllocateData.objects.create(agent_to=request.POST.get("agent_to"), product=request.POST.get("product"))
        messages.success(request, "Allocation Created Successfully")
        return redirect("allocate_data_list")
    return render(request, "allocate_data/create.html")

def allocate_data_delete(request, pk):
    get_object_or_404(AllocateData, pk=pk).delete()
    messages.success(request, "Allocation Deleted Successfully")
    return redirect("allocate_data_list")


# ─────────────────────────────────────────────────────────────────────────────
# Reallocate (web form + API)
# ─────────────────────────────────────────────────────────────────────────────

def reallocate_leads_view(request):
    if request.method == "POST":
        current_user = request.POST.get("current_assigned_user")
        reassign_to  = request.POST.get("reassign_to")
        leads = UploadData.objects.all() if current_user == "all_users" else UploadData.objects.filter(assigned_user=current_user)
        leads.update(assigned_user=reassign_to)
        ReallocateAssignedLead.objects.create(current_assigned_user=current_user, reassign_to=reassign_to)
        messages.success(request, "Leads Reallocated Successfully")
        return redirect("reallocate_leads")
    return render(request, "reallocate_leads/reallocate.html")


class ReallocateAssignedLeadAPIView(APIView):
    def post(self, request):
        serializer = ReallocateAssignedLeadSerializer(data=request.data)
        if serializer.is_valid():
            current_user = serializer.validated_data.get("current_assigned_user")
            reassign_to  = serializer.validated_data.get("reassign_to")
            leads = UploadData.objects.all() if current_user == "all_users" else UploadData.objects.filter(assigned_user=current_user)
            leads.update(assigned_user=reassign_to)
            serializer.save()
            return Response({"message": "Leads Reallocated Successfully"}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)